# preprocess excel file to add column labels
import os
import sys
import openpyxl
from openpyxl import load_workbook
import pandas as pd
import math
from pathlib import Path


if (len(sys.argv) != 2):
    raise ValueError("<usage>: python log_parser.py <path to excel file>")

file_path = sys.argv[1]
if (type(file_path) != str or file_path.strip() == ""):
    raise ValueError("Invalid path")

if (not os.path.isfile(file_path)):
    raise ValueError("File '" + file_path + "' does not exist")

# Done after to so it doesn't error out on a nonstring type
file_path = file_path.strip()
print("Fetching file " + file_path)

excel_file = file_path
# Load the existing Excel file
try:
    wb = load_workbook(excel_file)
except Exception as e:
    print(f"Error: {e}")
    sys.exit(1)


# Select the worksheet where you want to add the column headers
sheet = wb.active

# Get the headers from the first row of the worksheet
headers = [cell.value for cell in sheet[1]]

# Check if the headers are numerical from 0 to the last column
numerical_headers = all(str(i) == header for i, header in enumerate(headers))

if not numerical_headers:
    print("No numerical column headers found, adding numerical column headers")
    # created numbered column headers up to the last active column
    last_column = sheet.max_column

    column_headers = []

    for i in range(0, last_column+1):
        column_headers.append(str(i))

    # print(column_headers)

    # Insert a new row at the top of the worksheet and write the column headers
    sheet.insert_rows(1)
    for col, header in enumerate(column_headers, 1):
        sheet.cell(row=1, column=col, value=header)
else:
    print("Numerical Column headers found, no need to add headers")
# Save the modified Excel file
wb.save(excel_file)


# Read the Excel file into a Pandas DataFrame
df = pd.read_excel(file_path)
df = pd.DataFrame(df)
df.columns = df.columns.astype(str)


# Get all the different types of data
categories = df['9'].unique() # gets unique values of column '9'


# create separate dataframes for each type specified in the 9th column (column J from the raw csv file)
new_logframes = {}

for category in categories:
    # assigns a new dataframe to the key of category that contains only the rows from the original dataframe df where the value in the 9th column (column J) is equal to the current category.
    new_logframes[category] = df[df['9'] == category]

final_dataframes = {}
file_names = []

for category in new_logframes:
    col_names = ['Time']
    col = 10 # start after column 9

    # get all column names from new_logframes
    while col < 60:
    #    create new blank column in final_dataframes column based on the name from new_logframes
        col_name = new_logframes[category].iloc[0,col]
        if (type(col_name) == str):
            col_names.append(col_name)
        # go to next next column as MavLink organizes the data in pairs as column header, value
        col+=2


    # add category names into an array to be used as file names
    file_names.append(category)

    # drop column 9
    new_logframes[category] = new_logframes[category].drop(columns = ['9'])
    
    # drop all paramter columns
    col = 10
    while col < 60:
        new_logframes[category] = new_logframes[category].drop(str(col), axis=1)
        col+=2
    
    # drop all byte data
    new_logframes[category] = new_logframes[category].drop(new_logframes[category].columns[1:9], axis=1)

    # print(col_names)
    # print(len(col_names))

    # drop all rows past length of col_names
    new_logframes[category] = new_logframes[category].iloc[:, 0:len(col_names)]

    # rename column names
    new_logframes[category].columns = col_names


# Create new output directory

# Define the name of the new directory
new_directory_name = file_path + "-organized-UAV-Logs"

# Create the directory using Path.mkdir()
new_directory = Path(new_directory_name)
new_directory.mkdir(exist_ok=True)  # exist_ok=True allows it to continue if the directory already exists

print(f"Directory '{new_directory_name}' created successfully.")
# Export all new dataframes to excel files
for category in new_logframes:
    new_logframes[category].to_excel(new_directory_name + '/' + category + '.xlsx', index=False)